import requests
import json
import jsonpath
import openpyxl

def test_add_multiple_students():
    # API
    api_url = "http://thetestingworldapi.com/api/studentsDetails"
    f = open('C:/Users/hthai/Desktop/Task_API/AddNewStudent.json')

    # Type cast into json format
    json_request = json.loads(f.read())

    # Excel code
    vk = openpyxl.load_workbook('C:/Users/hthai/Desktop/Task_API/TestData.xlsx')
    sh = vk['Sheet1']
    rows = sh.max_row

    # starting a loop from second row
    for i in range(2, rows+1):
        # create 4 objects for each cell in every row
        cell_first_name = sh.cell(row = i, column = 1)
        cell_mid_name = sh.cell(row = i, column = 2 )
        cell_last_name = sh.cell(row = i, column = 3)
        cell_dob = sh.cell(row = i, column = 4)

        # Reading from each cell
        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_mid_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_dob.value

        # send the request to the server
        response = requests.post(api_url, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201
